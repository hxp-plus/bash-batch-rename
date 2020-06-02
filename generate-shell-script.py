import openpyxl
from openpyxl import load_workbook
import re
import sys

book_name='names.xlsx'
generated_script='rename.sh'
col_start=1
row_start=3

template="""
for file in `find . -regex '.*%s.*.docx'`; do
    mv -vn $file "%s-%s-%s.docx"
done
"""

remove_space="""
find $1 -name "* *.docx" -type f -print0 | \
  while read -d $'\\0' f; do mv -vn -v "$f" "${f// /_}"; done
"""

fo=open(generated_script,'w')

fo.write(remove_space)

workbook=openpyxl.load_workbook(book_name)
sheet=workbook.active

for i in range(row_start, 50):
    if(sheet.cell(row=i,column=col_start).value == None):
        break
    number=sheet.cell(row=i,column=col_start).value
    name=sheet.cell(row=i,column=col_start+1).value
    student_id=sheet.cell(row=i,column=col_start+2).value
    fo.write(template % (student_id[-5:],number,name,student_id))

